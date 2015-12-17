# anly_jiebaCutword.py
# -*- coding: utf8 -*-

import os
import jieba
import jieba.posseg as pseg
import jieba.analyse
import csv
from openpyxl import load_workbook
from openpyxl import Workbook
import sys
reload(sys)
sys.setdefaultencoding('utf8')


def fenci(infile,outfile,anlysis_clown):

    csvfile = file(infile, 'rb')
    reader = csv.reader(csvfile)
    writer = csv.writer(file(outfile,'wb'))
    srouce=''
    row = 0
    for line in reader:
        srouce+=''+line[anlysis_clown]
        row +=1
        # if row == 10:
        #     break
    print srouce.decode("gbk")

    # a=srouce.decode('string-escape').decode("gbk")
    a = srouce
    flag=[]
    fenci_list=[]
    dis_list = []
    fenci_list=list(jieba.cut(a, cut_all=False))

    for word in fenci_list:
        if word not in dis_list:
            dis_list.append(word)

    node_title=['word','count','flag']
    writer.writerow(node_title)

    for word in dis_list:
        count=fenci_list.count(word)
        flag=list(pseg.cut(word))[0].flag
        # print flag
        node_data=[word,count,flag]
        avilage_list=['n','ng','nr','nrt','ns','nt','nz']
        if node_data[2] in avilage_list:
            writer.writerow(node_data)
    csvfile.close()


if __name__=='__main__':

    # infile="E:/project/project-CAC/csv/new/test_fenci.csv"
    # outfile='E:/project/project-CAC/csv/new/test_fenci_result.csv'
    # anlysis_clown=0

    infile = 'D:\project\project-CAC\csv\cac.csv'
    outfile = 'D:\project\project-CAC\csv\cac_result.csv'
    fenci(infile,outfile,0)


'''
    # 读取excel文件
    lib = load_workbook('E:/project/project-CAC/csv/new/test_2.xlsx', use_iterators = True)
    # 创建新的excel文件
    wt_wb = Workbook(write_only=True)
    wt_ws = wt_wb.create_sheet()

    row_num=0
    data_all=''
    for row in lib['Sheet1'].iter_rows():
        row_num += 1
        # print row_num
        data=row[0].value
        data_all+='%s---'%data
        # print data
        seg_list = jieba.cut(data, cut_all=False)
        fenci_result="/ ".join(seg_list)
        # print fenci_result
        fenci_result_list=[data,fenci_result]
        textrank=jieba.analyse.textrank(data,3,False)
        extract_tags=jieba.analyse.extract_tags(data,3,True)

        # wt_ws.append(fenci_result_list)
        # print type(extract_tags)
        wt_ws.append([data,"-".join(textrank)])
    # 保存输出文件
    wt_wb.save('E:/project/project-CAC/csv/new/result_test.xlsx')



    textrank=jieba.analyse.textrank(data_all,30,False)
    # print textrank
    # help(jieba.analyse.textrank)
    # wt_ws.append(textrank)
    # wt_wb.save('E:/project-CAC/csv/new/result_test.xlsx')

    extract_tags=jieba.analyse.extract_tags(data_all,20,True)
    # print extract_tags
    # wt_ws.append(extract_tags)
    # wt_wb.save('E:/project-CAC/csv/new/result_test1.xlsx')

'''

