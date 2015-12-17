# con_mongo.py
import pymongo
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import os
# from openpyxl import load_workbook
from openpyxl import Workbook
os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'


# connect to mongoDB
client = pymongo.MongoClient('10.0.0.200', 27017)
db = client['test']
tables_list = db.collection_names()

# create excel file of  dataModel
wt_wb = Workbook(write_only=True)
wt_ws = wt_wb.create_sheet(title='NOTE')

for table in tables_list:
	table_name = table
	table_cnt= db[table].count()
	help(db[table])
	col = [table_name, table_cnt]
	wt_ws.append(col)



for table in tables_list:
# for table in ['item']:
	wt_ws1 = wt_wb.create_sheet(title=table)
	print table
	cursor = db[table].find_one()
	print cursor

	# for key, value in cursor.iteritems():
		# col_1 = [table, key]
		# print type(col_1)
		# wt_ws1.append(col_1)



wt_wb.save('D:\project\project-py\data\data_samply.xlsx')

# # open file
os.system('D:\project\project-py\data\data_samply.xlsx')