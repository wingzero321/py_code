# data_combinSheet.py
# -*- coding: utf8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook


# 读取excel文件
lib = load_workbook('E:/project-CAC/csv/new/survey2014.xlsx', use_iterators = True)

# E:\project-CAC\csv\new\
print lib.get_sheet_names()

# 创建新的excel文件
wt_wb = Workbook(write_only=True)
wt_ws = wt_wb.create_sheet()

# 将读取的记录复制到新的excel空文件中
sheetnames=lib.get_sheet_names()
for ws in lib.get_sheet_names():
	index =0
	for row in lib[ws].iter_rows():
		if ws != sheetnames[0] and index ==0:
			index +=1
			continue;
		write_row=[]
		for cell in row:
			write_row.append(cell.value)
		wt_ws.append(write_row)
		# print write_row

# 保存输出文件
wt_wb.save('E:/project-py/data/out_file.xlsx')

print 'end of all'