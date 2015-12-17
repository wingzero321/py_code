# ecc_cate_csv.py
# -*- coding: utf8 -*- 
import CatCon
import sys
import csv
import jieba
import jieba.posseg as pseg
import jieba.analyse
from openpyxl import load_workbook
from openpyxl import Workbook
import sys
reload(sys)
sys.setdefaultencoding('utf8')
# sys.setdefaultencoding('gb18030')
reload(sys)

print sys.getdefaultencoding()


import CCsettings as settings


class CCAgent:
    def __init__(self):
        self.cc_client = CatCon.CatConClient()
        self.cc_client.addServer(settings.cc_host,settings.cc_port)
        self.cc_client.usePersistentConnection(True)

    def getCCResult(self,message):
        cc_result = self.cc_client.categorize(message.encode('utf-8'),
                                            settings.article_type,
                                            "",
                                            settings.projects,
                                            settings.category_relevancy_type,
                                            0,
                                            settings.use_relevancy_cutoff,
                                            settings.cat_default_relevancy_cutoff,
                                            settings.skip_qualifier_matches)
        return cc_result


    def getCCResultTest(self,message,projects):
        cc_result = self.getCCResult(message)
        cate_result=''
        for cat in cc_result[projects]:
                cate_id=cat.getUniqueIDMetadata()
                Comments=cat.getCommentsMetadata()
                # cate_result+='%s##'%Comments
                cate_result=Comments        
        return cate_result



def getWordsCount(file,sheets_name,anlysis_row_nm,out_file):
    # 读取excel文件
    lib = load_workbook(file, use_iterators = True)
    # 创建新的excel文件
    wt_wb = Workbook(write_only=True)
    wt_ws = wt_wb.create_sheet()
    word_all=''
    fenci_list=[]
    dis_list = []
    words_sum=[]
    # 文档分词
    for row in lib[sheets_name].iter_rows():
        # print row[1].value
        fenci_list = list(jieba.cut(row[anlysis_row_nm].value, cut_all=False))
        fenci_list_p = list(pseg.cut(row[anlysis_row_nm].value))

        for w in fenci_list_p:
            word = [w.word,w.flag]
            # print word
            words_sum.append(word)
    #生成字典    
    for word in words_sum:
        if word not in dis_list:
            dis_list.append(word)
    #计算词频
    wt_ws.append(['words','notes','count'])
    for word in dis_list:
        count = words_sum.count(word)
        word_c = [word[0],word[1],count]
        wt_ws.append(word_c)
    # 保存excel
    wt_wb.save(out_file)



if __name__=='__main__':

    # getWordsCount(file,anlysis_row_nm,out_file)

    # getWordsCount(
    #     'D:/project/project-QVOC2/csv/data/PCRs/ExportFile_PDI.xlsx',
    #     'PDI',
    #     6,
    #     'D:/project/project-QVOC2/csv/data/PCRs/result_PDI_wordcount.xlsx')

    getWordsCount(
        'D:/project/project-QVOC2/csv/data/PCRs/ExportFile_PT.xlsx',
        'Sheets1',
        6,
        'D:/project/project-QVOC2/csv/data/PCRs/ExportFile_PT_wordcount.xlsx')



    # 读取excel文件
    # lib = load_workbook('E:/project-CAC/csv/new/test.xlsx', use_iterators = True)
    # lib = load_workbook('D:/project/project-QVOC2/csv/data/PCRs/ExportFile_PDI.xlsx', use_iterators = True)
    # # 创建新的excel文件
    # wt_wb = Workbook(write_only=True)
    # wt_ws = wt_wb.create_sheet()

    # cc_agent = CCAgent()
    # row_num=0
    # for row in lib['PDI'].iter_rows():
    #     # print row[anlysis_row_nm].value
    #     row_num+=1
    #     print row_num
    #     data=row[anlysis_row_nm].value
    #     cc_result = cc_agent.getCCResultTest(data,'bbs_qvoc')
    #     write_row=[data,cc_result]
    #     wt_ws.append(write_row)
    #     # if row_num>=100:
    #         # break;
    # # 保存输出文件
    # wt_wb.save('D:/project/project-QVOC2/csv/data/PCRs/ExportFile_PDI_result.xlsx')
    