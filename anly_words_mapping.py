# anly_words_mapping.py
# -*- coding: utf8 -*- 
import sys
import jieba
import json
# import jieba.posseg as pseg
import jieba.analyse
from openpyxl import load_workbook
from openpyxl import Workbook
reload(sys)
sys.setdefaultencoding('utf8')
reload(sys)


# 创建词向量
def createVocabLst(dataSet):
	vocabSet = set([])
	for document in dataSet:
		vocabSet = vocabSet | set(document)
	return list(vocabSet)


# 计算夹角余余弦
def cosVector(x,y):
    if(len(x)!=len(y)):
        print('error input,x and y is not in the same space')
        return;
    result1=0.0;
    result2=0.0;
    result3=0.0;
    for i in range(len(x)):
        result1+=x[i]*y[i]   #sum(X*Y)
        result2+=x[i]**2     #sum(X*X)
        result3+=y[i]**2     #sum(Y*Y)
    # print("result is "+str(result1/((result2*result3)**0.5))) #结果显示
    cosVector = result1/((result2*result3)**0.5)
    return cosVector

# 词向量转化为数值向量
def setOfWords2Vec(vocabList,inputSet):
	returnVec = [0]*len(vocabList)
	for word in inputSet:
		if word in vocabList:
			returnVec[vocabList.index(word)] = 1
		else :print "the word: %s is not in my Vocabulary!" % word
	return returnVec


# 中文打印
def printChinese(List):
	str = json.dumps(List,ensure_ascii=False)
	return str


def test_tfidf():
	from sklearn.feature_extraction.text import TfidfTransformer
	from sklearn.feature_extraction.text import CountVectorizer

	if __name__ == "__main__":
	  corpus=["我 来到 北京 清华大学",#第一类文本切词后的结果，词之间以空格隔开
	    "他 来到 了 网易 杭研 大厦",#第二类文本的切词结果
	    "小明 硕士 毕业 与 中国 科学院",#第三类文本的切词结果
	    "我 爱 北京 天安门"]#第四类文本的切词结果
	  vectorizer=CountVectorizer()#该类会将文本中的词语转换为词频矩阵，矩阵元素a[i][j] 表示j词在i类文本下的词频
	  transformer=TfidfTransformer()#该类会统计每个词语的tf-idf权值
	  tfidf=transformer.fit_transform(vectorizer.fit_transform(corpus))#第一个fit_transform是计算tf-idf，第二个fit_transform是将文本转为词频矩阵
	  word=vectorizer.get_feature_names()#获取词袋模型中的所有词语
	  weight=tfidf.toarray()#将tf-idf矩阵抽取出来，元素a[i][j]表示j词在i类文本中的tf-idf权重
	  for i in range(len(weight)):#打印每类文本的tf-idf词语权重，第一个for遍历所有文本，第二个for便利某一类文本下的词语权重
	    print u"-------这里输出第",i,u"类文本的词语tf-idf权重------"
	    for j in range(len(word)):
	      print word[j],weight[i][j]




def main():

	lib = load_workbook('D:\project\project-carNew\csv\dealer_name_mapping.xlsx')
	# 创建新的excel文件
	out_file = 'D:\project\project-carNew\csv\dealer_name_mapping_result.xlsx'

	wt_wb = Workbook(write_only=True)
	wt_ws = wt_wb.create_sheet()

	wordsList = []
	std_wordsList = []
	all_wordsList = []

	for row in lib['Sheet1'].iter_rows():
		fenci_list = list(jieba.cut(row[0].value, cut_all=False))
		wordsList.append(fenci_list)
		all_wordsList.append(fenci_list)


	for row in lib['Sheet2'].iter_rows():
		fenci_list = list(jieba.cut(row[0].value, cut_all=False))
		std_wordsList.append(fenci_list)
		all_wordsList.append(fenci_list)


	print printChinese(all_wordsList)

	myVocabLst = createVocabLst(all_wordsList)

	row_nm=0
	for words in wordsList:
		words_vec = setOfWords2Vec(myVocabLst, words)
		# print words_vec

		n=0
		right_word= []
		for std_words in std_wordsList:
			std_word_vec = setOfWords2Vec(myVocabLst, std_words)
			cos_nm = cosVector(words_vec,std_word_vec)
			# print cos_nm
			if cos_nm>n:
				print cos_nm
				n = cos_nm
				right_word = std_words

		left_word = ''.join(words)
		ok_word = ''.join(right_word)

		writeRowList = [left_word,ok_word,n]
		print printChinese(writeRowList)
		if row_nm == 0:
			wt_ws.append(['left_word','ok_word','cos_nm'])
		else:
			wt_ws.append(writeRowList)
		row_nm+=1
		# if row_nm == 10:break


	wt_wb.save(out_file)

if __name__=='__main__':


	main()
	# test_tfidf()


