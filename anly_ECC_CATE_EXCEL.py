# anly_ECC_CATE_EXCEL.py
# -*- coding: utf8 -*- 
import CatCon
import sys
import csv
import jieba
import jieba.posseg as pseg
import jieba.analyse
from openpyxl import load_workbook
from openpyxl import Workbook
import sys
reload(sys)
sys.setdefaultencoding('utf8')
# sys.setdefaultencoding('gb18030')
reload(sys)

print sys.getdefaultencoding()


import CCsettings as settings


class CCAgent:
    def __init__(self):
        self.cc_client = CatCon.CatConClient()
        self.cc_client.addServer(settings.cc_host,settings.cc_port)
        self.cc_client.usePersistentConnection(True)

    def getCCResult(self,message):
        cc_result = self.cc_client.categorize(message.encode('utf-8'),
                                            settings.article_type,
                                            "",
                                            settings.projects,
                                            settings.category_relevancy_type,
                                            0,
                                            settings.use_relevancy_cutoff,
                                            settings.cat_default_relevancy_cutoff,
                                            settings.skip_qualifier_matches)
        return cc_result


    def getCCResultTest(self,message,projects):
        cc_result = self.getCCResult(message)
        cate_result=''
        for cat in cc_result[projects]:
                # cate_id=cat.getUniqueIDMetadata()
                Comments=cat.getCommentsMetadata()
                cate_result+='%s##'%Comments
                # cate_result=Comments        
        return cate_result



def getWordsCount(file,sheets_name,anlysis_row_nm,out_file):
    # 读取excel文件
    lib = load_workbook(file, use_iterators = True)
    # 创建新的excel文件
    wt_wb = Workbook(write_only=True)
    wt_ws = wt_wb.create_sheet()
    word_all=''
    fenci_list=[]
    dis_list = []
    words_sum=[]
    # 文档分词
    for row in lib[sheets_name].iter_rows():
        # print row[1].value
        fenci_list = list(jieba.cut(row[anlysis_row_nm].value, cut_all=False))
        fenci_list_p = list(pseg.cut(row[anlysis_row_nm].value))

        for w in fenci_list_p:
            word = [w.word,w.flag]
            # print word
            words_sum.append(word)
    #生成字典    
    for word in words_sum:
        if word not in dis_list:
            dis_list.append(word)
    #计算词频
    wt_ws.append(['words','notes','count'])
    for word in dis_list:
        count = words_sum.count(word)
        word_c = [word[0],word[1],count]
        wt_ws.append(word_c)
    # 保存excel
    wt_wb.save(out_file)


def getCate_csv():
    cc_agent = CCAgent()
    infile = 'D:\project\project-iproduct\csv\POC2/auto_koubei.csv'
    outfile = 'D:\project\project-iproduct\csv\POC2/auto_koubei_test.csv'
    writer = csv.writer(file(outfile,'wb'))
    csvfile = file(infile, 'rb')
    reader = csv.reader(csvfile)
    row_num=0
    for line in reader:
        
        # data = line[1].decode('string-escape').decode("gbk")
        try:
            data = line[1].decode('string-escape').decode("gbk")
            cc_result = cc_agent.getCCResultTest(data,'UGC_keypoint')
        except:
            data = ''
            cc_result = ''

        print row_num
        write_row=[data,cc_result]
        writer.writerow(write_row)

        row_num+=1
        # if row_num>=100:
        #     break;


def GetCategories():

    ccClient = CatCon.CatConClient()
    ccClient.addServer(settings.cc_host,settings.cc_port)

    avail_projects = ccClient.listAvailableCategorizationProjects()

    print avail_projects

    categories = ccClient.listAvailableCategoriesWithMetadata([avail_projects[0]])
    categorie = categories[avail_projects[0]]

    print categorie[10]
    print len(categorie)

    FIELDS = categorie[10].keys()
    # print FIELDS
    csv_file = open('D:\project\project-iproduct\csv\list_ecc_projects.csv', 'wb')
    writer = csv.DictWriter(csv_file, fieldnames=FIELDS)
    writer.writerow(dict(zip(FIELDS, FIELDS)))
    # d = categorie[10]
    for project in avail_projects:
        print project
        categories = ccClient.listAvailableCategoriesWithMetadata([project])
        categorie = categories[project]

        for item in categorie:
            writer.writerow(item)
    
    csv_file.close()


if __name__=='__main__':

    GetCategories()

    # getCate_csv()

    # 读取excel文件
    # lib = load_workbook('E:/project-CAC/csv/new/test.xlsx', use_iterators = True)
    # lib = load_workbook('D:\project\project-iproduct\csv\POC2/test.xlsx',use_iterators = True)
    # # 创建新的excel文件
    # wt_wb = Workbook(write_only=True)
    # wt_ws = wt_wb.create_sheet()
    # anlysis_row_nm = 1
    # cc_agent = CCAgent()
    # row_num=0
    # for row in lib['Sheet1'].iter_rows():
    #     # print row[anlysis_row_nm].value
    #     row_num+=1
    #     print row_num
    #     data=row[anlysis_row_nm].value
    #     cc_result = cc_agent.getCCResultTest(data,'Config_expectation')
    #     write_row=[data,cc_result]
    #     wt_ws.append(write_row)
    #     if row_num>=100:
    #         break;
    # # 保存输出文件
    # wt_wb.save('D:\project\project-iproduct\csv\POC2/test_result.xlsx')
