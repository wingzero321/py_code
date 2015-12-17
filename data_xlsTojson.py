# data_xlsTojson.py
# -*- coding: utf8 -*- 


from openpyxl import load_workbook
from openpyxl import Workbook
from json import *
import os
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

# 读取excel文件

file = 'D:/project/project-py/data/best.xlsx'

lib = load_workbook(file, use_iterators = True)
sheetnames = lib.get_sheet_names()
ws = lib.get_sheet_by_name(sheetnames[0])
print "file is ", file
print   "Work Sheet Titile:" ,ws.title
print   "Work Sheet Rows:" ,ws.get_highest_row() 
print   "Work Sheet Cols:" ,ws.get_highest_column()

import json
# import simplejson
# 拼接成json格式
data = ''
obs = 0
print dir(json)
for row in lib[sheetnames[0]].iter_rows():
	obs+=1
	# print obs
	dict={}
	dict['name']=row[0].value
	dict['value']=row[1].value
	dict['itemStyle']="createRandomItemStyle()"
	# print row[0].value
	# js=JSONEncoder().encode(dict)
	js= json.dumps(dict,ensure_ascii=False)
	# js=str(dict)
	# print row[0].value.encode('gbk').decode('gbk').encode('utf-8')
	# print dict
	if obs==2:
		data=js
	if obs>2:
		data+=","+js
	print js
	# print data.encode('gb18030')


print data

# 文件输出
# wt_wb = Workbook(write_only=True)
# wt_ws = wt_wb.create_sheet(title='NOTE')
# wt_ws.append([data])
# wt_wb.save('D:/project/project-py/data/output_wc.xlsx')

testDict = {'a':1,'b':2}

file = 'D:\project\project-py\data\my.json'
fp = open(file,'w+')
fp.write(data)
fp.close()