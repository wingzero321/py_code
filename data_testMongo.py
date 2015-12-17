# data_testMongo.py
# -*- coding: utf8 -*- 

import pymongo
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import os
from openpyxl import load_workbook
from openpyxl import Workbook


def dictToString(dict):
	str= ''
	for item in dict:
		str += item + "//"
	return str


def db_info(db_name):
	wt_ws = wt_wb.create_sheet(title=db_name)
	# query databases info
	table_list = mongo_client[db_name].collection_names()
	table_ex=[u'fs.files',u'fs.chunks',u'system.indexes']
	table_list_a = []
	for table in table_list:
		if table not in table_ex:
			table_list_a.append(table)
	print table_list_a

	inert_data=[]  # query info data
	title = ['table_name','count','column_count','column_str']
	wt_ws.append(title)
	for collection in table_list_a:
		print '>>>>>>>>>>',collection,'<<<<<<<<<<<'
		col = mongo_client[db_name][collection]
		data = col.find_one()
		# print data
		# 
		if data != None:
			column_c = len(data)
			column=data.keys()
			column_str = dictToString(data.keys())
		else :
			column_c=0
			column=[]
			column_str= ''

		print column
		info =  [collection,col.find().count(),column_c,column_str]
		print info
		inert_data.append(info)

	# inert_data.sort(key=lambda x:x[1])
	# print inert_data
	

	for item in inert_data:
		wt_ws.append(item)


if __name__=='__main__':

	# setting
	mongo_host = "10.0.0.200"
	mongo_port = 27017
	out_file = 'D:\project\project-py\data\data_table_info.xlsx'
	# connect to mongoDB
	mongo_client = pymongo.MongoClient( mongo_host, mongo_port ,tz_aware = True )


	# create info excel file
	wt_wb = Workbook(write_only=True)

	db_info('test')
	# db_info('social_survey')


	# save output file
	wt_wb.save(out_file)
	# os.system(out_file)