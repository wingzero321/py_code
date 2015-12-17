# anly_textMiner_forIproduct.py
# -*- coding: utf8 -*-


import csv
import re
from snownlp import SnowNLP
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import sys
reload(sys)
sys.setdefaultencoding('utf8')
import jieba
import jieba.posseg as pseg
import jieba.analyse


def getWordsCount(in_file,out_file):
    # 读取excel文件
    # lib = load_workbook(file, use_iterators = True)
    # 创建新的excel文件
    wt_wb = Workbook(write_only=True)
    wt_ws = wt_wb.create_sheet()
    word_all=''
    fenci_list=[]
    dis_list = []
    words_sum=[]

    lib = csv.reader(file(in_file, 'rb'))
    # 文档分词
    for row in lib:
        fenci_list_p = list(pseg.cut(row[0]))

        for w in fenci_list_p:
            word = [w.word,w.flag]
            # print word
            words_sum.append(word)
    #生成字典    
    for word in words_sum:
        if word not in dis_list:
            dis_list.append(word)
    #计算词频
    wt_ws.append(['words','notes','count'])
    for word in dis_list:
        count = words_sum.count(word)
        word_c = [word[0],word[1],count]
        wt_ws.append(word_c)
    # 保存excel
    wt_wb.save(out_file)


def getKeyword(key_word,split_word,ss):

	s = ss
	# ss = s.split(split_word)
	ss = re.split(u'。|？|！|】|，|,|   |',s)
	# print len(ss)
	words_list = [x.lower() for x in key_word.split()]	
	out_words=''
	for item in ss:
		if all([word in item.lower() and True or False for word in words_list]):
			out_words+=item+"///"

	if len(out_words)<=14:
		ss = re.split(u'。|？|！|【',s)
		words_list = [x.lower() for x in key_word.split()]	
		out_words=''
		for item in ss:
			if all([word in item.lower() and True or False for word in words_list]):
				out_words+=item+"///"
	
	print '>>>><<<<<<',len(out_words)
	return out_words


def getKeywordforFile(in_file,column_num,words,out_file,out_wc_file):
	lib = load_workbook(in_file, use_iterators = True)
	num=0
	writer = csv.writer(file(out_file, 'wb'))
	for row in lib[words].iter_rows():
		num+=1
		print num
		out_words=getKeyword(words,"\\n",row[column_num].value)

		if out_words=='':
			out_words=getKeyword('ACC',"\\n",row[column_num].value)


		# 情感分析
		Positives = [u'好',u'不错',u'很好',u'实用',u'棒',u'加装',u'先进',u'喜欢',u'操作简单',u'很爽',u'当然',u'开启',u'加强',u'有用',u'舒适',u'方便',u'高科技',u'强大',u'给力',u'炫',u'人性化',u'稳定',u'无压力',u'神器',u'赞',u'顺手',
		u'满意',u'省心',u'轻松',u'省体力',u'优势',u'优点',u'不得不提',u'享受',u'配置了',u'很帅',u'满足',u'上档次',u'自动亮',u'吸引',u'骚到不行',u'也装',u'漂亮',u'厚道',u'加入了',u'屌',u'犀利',u'吸引',u'特别',u'亮点',u'美',
		u'档次',u'爽',u'大气',u'安全',u'个性',u'霸气',u'美感',u'档次',u'亮眼',u'帅',u'少有',u'时髦',u'卖点',u'采用',u'省事',u'配置全有',u'性价比高',u'高大上',u'多了',u'丰富',u'装逼',u'高大上',u'一应俱全',u'大爱',u'最爱']

		Negatives = [u'不好',u'算了',u'没用过',u'才行',u'不喜欢',u'不爽',u'神经',u'无法',u'报错',u'不到位',u'用处不大',u'普通',u'找不到',u'有缝隙',u'损坏',u'不行',u'不够',u'落伍',u'不值',u'不是四门全开',u'不管用',u'不太灵敏']

		# Wanteds = [u'没有',u'无',u'尤其',u'完美',u'加配',u'减配',u'缺',u'少了',u'流行',u'应该标配',u'选装',u'高端',u'自购',u'标配',u'虽然',u'将配置',u'简配',u'就为',u'换了',u'没钱',u'更换',u'换',u'不是led',u'羡慕',u'建议']
		Wanteds = [u'没有',u'尤其',u'完美',u'加配',u'减配',u'缺',u'少了',u'流行',u'应该标配',u'选装',u'高端',u'自购',u'标配',u'虽然',u'将配置',u'简配',u'就为',u'换了',u'没钱',u'更换',u'换',u'不是led',u'羡慕',u'建议',u'都有',u'高配上才有',u'为什么不加上',u'最理想']

		cat_lists = [Positives,Negatives,Wanteds]
		setmention=''
		keys=''
		for key in Wanteds:
			if re.findall(key, out_words) != []:
				setmention='Wanted'
				keys=key
		for key in Positives:
			if re.findall(key, out_words) != []:
				setmention='Positive'
				keys=key

		for key in Negatives:
			if re.findall(key, out_words) != []:
				setmention='Negative'
				keys=key

		if setmention == '':
			setmention='NA'

		# print out_words,setmention
		writer.writerow([row[column_num].value.encode('gb18030'),out_words.encode('gb18030'),words.encode('gb18030'),setmention,keys.encode('gb18030')])
		# fenci
		# getWordsCount(out_file,out_wc_file)


if __name__=='__main__':


	# getKeywordforFile(
	# 	'D:\project\project-iproduct\csv\poc_pinglun.xlsx',
	# 	0,
	# 	u'自适应巡航',
	# 	'D:/project/project-iproduct/csv/test_out_words.csv',
	# 	'D:/project/project-iproduct/csv/poc_pinglun_wc_acc.xlsx'
	# 	)


	# getKeywordforFile(
	# 	'D:\project\project-iproduct\csv\poc_pinglun.xlsx',
	# 	0,
	# 	u'LED',
	# 	'D:/project/project-iproduct/csv/test_out_words_led.csv',
	# 	'D:/project/project-iproduct/csv/poc_pinglun_wc_led.xlsx'
	# 	)

	getKeywordforFile(
		'D:\project\project-iproduct\csv\poc_pinglun.xlsx',
		0,
		u'无钥匙进入',
		'D:/project/project-iproduct/csv/test_out_words_nk.csv',
		'D:/project/project-iproduct/csv/poc_pinglun_wc_nk.xlsx'
		)

	# lib = load_workbook('D:\project\project-iproduct\csv\poc_pinglun.xlsx', use_iterators = True)
	# writer = csv.writer(file('D:/project/project-iproduct/csv/test_out_words.csv', 'wb'))
	

	
	# 提取关键字所在的句子
	# words=u'LED'
	# for row in lib[words].iter_rows():
	# 	out_words=getKeyword(u'LED',"\\n",row[0].value)
	# 	print out_words
	# 	writer.writerow([row[0].value.encode('gb18030'),out_words.encode('gb18030'),words])


	# getWordsCount(
	# 	'D:\project\project-iproduct\csv\poc_pinglun.xlsx',
	# 	'LED',
	# 	2,
	# 	'D:\project\project-iproduct\csv\poc_pinglun_wc_led.xlsx')



	
	
	# 提取关键字所在的句子
	# words=u'无钥匙进入'
	# out_file = 'D:/project/project-iproduct/csv/test_out_words.csv'
	# num=0
	# for row in lib[words].iter_rows():
	# 	num+=1
	# 	print num
	# 	out_words=getKeyword(words,"\\n",row[0].value)
	# 	print out_words
	# 	writer = csv.writer(file(out_file, 'wb'))
	# 	writer.writerow([row[0].value.encode('gb18030'),out_words.encode('gb18030'),words.encode('gb18030')])




# csv key提取与情感分析
# csvfile = file('D:/project/project-iproduct/csv/getPublicPraiseData2.csv', 'rb')
# csvfile = file('D:/project/project-iproduct/csv/test.csv', 'rb')
# writer = csv.writer(file('D:/project/project-iproduct/csv/test_out.csv', 'wb'))
# reader = csv.reader(csvfile)

# for line in reader:
# 	key_words = u'车道保持'
# 	s = line[0].decode('gb18030')
# 	ss = s.split("\\n")
# 	print len(ss)
# 	words_list = [x.lower() for x in key_words.split()]	
# 	out_words=''
# 	for item in ss:
# 	    if all([word in item.lower() and True or False for word in words_list]):
# 	    	out_words+=item+"///"
# 	print out_words

# 	sentiment=SnowNLP(s).sentiments
# 	writer.writerow([key_words.encode('gb18030'),s.encode('gb18030'), out_words.encode('gb18030'),sentiment])


# row=0
# for line in reader:
# 	row+=1

# 	if row == 1:
# 		n=0
# 		for column in line:
# 			print n,column
# 			n+=1

# 	key_words = 'ESP'
# 	s = line[18]
# 	ss = s.split("\\n")
# 	print len(ss)
# 	words_list = [x.lower() for x in key_words.split()]

# 	out_words=''
# 	for item in ss:
# 	    if all([word in item.lower() and True or False for word in words_list]):
# 	    	out_words+='///'+item
# 	        # print item

# 	s1 = out_words.decode('gb18030')
# 	s2 = SnowNLP(s1)
# 	print s1.sentiments

# 	if row>100:
# 		break