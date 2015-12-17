# data_combine_xls.py
# -*- coding: utf8 -*- 


import os
from openpyxl import load_workbook
from openpyxl import Workbook
import xlrd



file_list=[]

def fun( path ):
	for root, dirs, files in os.walk( path ):
		for fn in files:
			# print root, fn
			file_list.append(root+'/'+fn)

# fun('D:\project\project-CA\csv\CP')
fun('D:\project\project-CA\csv\Inventory_report')
file = file_list[0]
file = file.replace("\\","/")
print file

# lib = load_workbook(file, use_iterators = True)

# # 取第一张表
# sheetnames = lib.get_sheet_names()
# ws = lib.get_sheet_by_name(sheetnames[0])  
# # #显示表名，表行数，表列数   
# print   "Work Sheet Titile:" ,ws.title  
# print   "Work Sheet Rows:" ,ws.get_highest_row()  
# print   "Work Sheet Cols:" ,ws.get_highest_column()  

# 创建新的excel文件
wt_wb = Workbook(write_only=True)
wt_ws = wt_wb.create_sheet()


for file in file_list:
	file = file.replace("\\","/")
	# print file
	data = xlrd.open_workbook(file)
	table = data.sheet_by_index(0)
	nrows = table.nrows
	print file,nrows
	for i in range(nrows):
		row_list = table.row_values(i)
		wt_ws.append(row_list)

wt_wb.save('D:\project\project-CA\csv\Inventy_result_test.xlsx')

		



	# table = data.sheet_by_index(0)
	# nrows = table.nrows
	# print file,nrows
	# for i in range(nrows):
 #      print table.row_values(i)
    # ncols = table.ncols


	# lib = load_workbook(file, use_iterators = True)
	# sheetnames = lib.get_sheet_names()
	# ws = lib.get_sheet_by_name(sheetnames[0])
	# print "file is ", file
	# print   "Work Sheet Titile:" ,ws.title
	# print   "Work Sheet Rows:" ,ws.get_highest_row() 
	# print   "Work Sheet Cols:" ,ws.get_highest_column()

	# for row in lib[sheetnames[0]].iter_rows():
		# for row_sub in row:
			# if row_sub.value = 'season':
			# print type(row_sub.value)
	# print row
